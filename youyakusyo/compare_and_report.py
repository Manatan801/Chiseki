"""
登記事項要約書 照合レポート生成
照合元CSV（調査開始時点）× OCR結果CSV（最新）→ Excel 3シート
"""

import csv
import re
import unicodedata
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# === 設定 ===
SOURCE_CSV_DIR = Path("/home/maaatan/Chiseki/youyakusyo/csv")
OCR_CSV = Path("/home/maaatan/Chiseki/youyakusyo/output/要約書_全データ.csv")
OUTPUT_XLSX = Path("/home/maaatan/Chiseki/youyakusyo/output/照合レポート.xlsx")

# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
CHANGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NEW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def zen_to_han(text: str) -> str:
    """全角数字・英字・記号を半角に変換"""
    return unicodedata.normalize("NFKC", text)


def normalize_chiban(raw: str) -> str:
    """地番を正規化（全角→半角、ハイフン統一、「番」統一）"""
    s = zen_to_han(raw).strip()
    s = s.replace("－", "-").replace("ー", "-").replace("—", "-")
    # ハイフン区切り → 「番」（例: 2325-1 → 2325番1）
    s = re.sub(r"(\d+)-(\d+)", r"\1番\2", s)
    # 「番」がなく純粋な数字のみの場合 → 末尾に「番」は付けない（そのまま）
    # 「882番8」と「882-8」は統一済み
    # 「番」を除去して数字のみで比較するキーも用意
    return s


def chiban_match_key(raw: str) -> str:
    """地番をマッチング用に正規化（「番」を統一的に処理）"""
    s = normalize_chiban(raw)
    # 「番の2」→「番2」（「の」除去）
    s = s.replace("番の", "番")
    # 末尾の「番」だけの場合を除去（例: 「895番」→「895」）
    s = re.sub(r"番$", "", s)
    return s


def normalize_text(text: str) -> str:
    """テキスト正規化（比較用）"""
    s = zen_to_han(text).strip()
    s = s.replace("　", " ").replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    # 「北茨城」の文字化け対応
    s = s.replace("北\u8328城", "北茨城")
    return s


# 登記情報提供サービスの外字コード → Unicode文字マッピング
GAIJI_MAP = {
    "<00008328>": "茨",
    "<00009AD9>": "髙",
    "<0000FA11>": "﨑",
    "<0000FA19>": "神",
    "<0000F9DC>": "隆",
    "<00005BEC>": "寬",
    "<00005EE3>": "廣",
    "<000080D6>": "胖",
    "<00008FC4>": "迄",
    "<000049FA>": "䧺",
    "<0000E452>": "〓",  # 私用領域（特定不能）
    "<00100F20>": "〓",  # 範囲外（特定不能）
    "<000F61BE>": "〓",  # 補助漢字（特定不能）
}


def replace_gaiji(text: str) -> str:
    """外字コード <XXXXXXXX> を実際の文字に置換"""
    for code, char in GAIJI_MAP.items():
        text = text.replace(code, char)
    # 未知の外字コードも処理
    def _replace_unknown(m):
        try:
            cp = int(m.group(1), 16)
            if 0 < cp < 0x110000:
                return chr(cp)
        except (ValueError, OverflowError):
            pass
        return "〓"
    text = re.sub(r"<([0-9A-Fa-f]{8})>", _replace_unknown, text)
    return text


def parse_source_csvs(csv_dir: Path) -> dict:
    """照合元CSV（Shift-JIS階層構造）をパースして物件辞書を返す"""
    properties = {}

    for csv_path in sorted(csv_dir.glob("*_001_BM.csv")):
        with open(csv_path, "r", encoding="cp932", errors="replace") as f:
            reader = csv.reader(f)
            current_prop = None
            current_num = None

            for row in reader:
                if len(row) < 3:
                    continue

                # 全セルに外字変換を適用
                row = [replace_gaiji(cell) for cell in row]

                num = zen_to_han(row[0]).strip()
                record_type = zen_to_han(row[1]).strip()

                if record_type == "物件情報":
                    # 新しい物件開始
                    raw_chiban = zen_to_han(row[5]).strip() if len(row) > 5 else ""
                    shozai = normalize_text(row[4]) if len(row) > 4 else ""
                    aza = normalize_text(row[9]) if len(row) > 9 else ""

                    current_num = num
                    current_prop = {
                        "source_file": csv_path.name,
                        "raw_chiban": raw_chiban,
                        "shozai": shozai,
                        "aza": aza,
                        "chimoku": "",
                        "chiseki": "",
                        "chimoku_old_values": [],  # 旧地目リスト（【】付きだったもの）
                        "owners": [],
                        "history": [],
                    }

                elif record_type.startswith("所在") and current_prop:
                    # 所在の詳細（字名含む）
                    full_shozai = normalize_text(row[2]) if len(row) > 2 else ""
                    if full_shozai:
                        current_prop["shozai_full"] = full_shozai

                elif record_type.startswith("表示履歴") and current_prop:
                    chiban_val = zen_to_han(row[2]).strip() if len(row) > 2 else ""
                    chimoku_raw = zen_to_han(row[4]).strip() if len(row) > 4 else ""
                    chiseki_raw = zen_to_han(row[5]).strip() if len(row) > 5 else ""

                    if chiban_val and not current_prop.get("chiban_display"):
                        current_prop["chiban_display"] = chiban_val

                    # 【】付き = 旧値（変更前）→ 無視
                    # 【】なし = 現在の値 → 採用
                    is_old_chimoku = bool(re.search(r"[【】]", chimoku_raw))
                    is_old_chiseki = bool(re.search(r"[【】]", chiseki_raw))

                    chimoku_clean = re.sub(r"[【】\[\]]", "", chimoku_raw).strip()
                    chiseki_clean = re.sub(r"[【】\[\]]", "", chiseki_raw).strip()

                    # 現在値のみ採用（旧値は上書きしない）
                    if chimoku_clean and not is_old_chimoku:
                        current_prop["chimoku"] = chimoku_clean
                    elif chimoku_clean and is_old_chimoku:
                        # 旧値をリストに保存（OCR誤読検出に使用）
                        current_prop["chimoku_old_values"].append(chimoku_clean)
                        if not current_prop["chimoku"]:
                            current_prop["chimoku_old"] = chimoku_clean

                    if chiseki_clean and not is_old_chiseki:
                        current_prop["chiseki"] = chiseki_clean
                    elif chiseki_clean and is_old_chiseki and not current_prop["chiseki"]:
                        current_prop["chiseki_old"] = chiseki_clean

                    current_prop["history"].append({
                        "chimoku": chimoku_raw, "chiseki": chiseki_raw,
                        "detail": zen_to_han(row[6]).strip() if len(row) > 6 else "",
                    })

                elif record_type.startswith("所有権") and current_prop:
                    addr = normalize_text(row[2]) if len(row) > 2 else ""
                    name = normalize_text(row[4]) if len(row) > 4 else ""
                    reg_info = normalize_text(row[6]) if len(row) > 6 else ""
                    ref_num = normalize_text(row[7]) if len(row) > 7 else ""

                    if name or addr:
                        current_prop["owners"].append({
                            "address": addr,
                            "name": name,
                            "reg_info": reg_info,
                            "ref_num": ref_num,
                        })

                # 次の物件に移る前に現在の物件を保存
                if record_type == "物件情報" and current_prop:
                    # 地番をキーにして保存
                    chiban_key = current_prop.get("chiban_display", current_prop["raw_chiban"])
                    if chiban_key:
                        norm_key = chiban_match_key(chiban_key)
                        properties[norm_key] = current_prop

            # 最後の物件を保存
            if current_prop:
                chiban_key = current_prop.get("chiban_display", current_prop["raw_chiban"])
                if chiban_key:
                    norm_key = chiban_match_key(chiban_key)
                    properties[norm_key] = current_prop

    return properties


def parse_ocr_csv(csv_path: Path) -> dict:
    """OCR結果CSVをパースして物件辞書を返す"""
    properties = {}

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            chiban_raw = (row.get("地番") or "").strip()
            if not chiban_raw:
                continue

            # 不正な地番をフィルタ（続きページで住所が地番に入るケース）
            # 地番は数字で始まるはず
            chiban_check = zen_to_han(chiban_raw)
            if not re.match(r"^\d", chiban_check):
                continue
            # 「番地」を含む場合は住所の誤認
            if "番地" in chiban_raw:
                continue

            norm_key = chiban_match_key(chiban_raw)

            if norm_key not in properties:
                properties[norm_key] = {
                    "shozai": (row.get("所在") or "").strip(),
                    "chiban_display": chiban_raw,
                    "chimoku": (row.get("地目") or "").strip(),
                    "chiseki": (row.get("地積（㎡）") or "").strip(),
                    "owners": [],
                    "page": (row.get("ページ") or ""),
                    "pdf_file": (row.get("PDFファイル") or ""),
                }

            # 所有者追加（属性ベースのカラム補正付き）
            owner_addr = (row.get("所有者・共有者住所") or "").strip()
            mochibun = (row.get("持分") or "").strip()
            owner_name = (row.get("所有者・共有者氏名") or "").strip()
            reg_date = (row.get("登記日") or "").strip()
            ref_num = (row.get("受付番号") or "").strip()

            # カラムずれ検出・補正
            # 正しいパターン（共有地）: 住所, 持分(持分XX分のX), 氏名, 日付(XX年), 受付(第XX号)
            # ずれパターン（単独所有）: 住所, 氏名→持分欄, 日付→氏名欄, 受付→日付欄, (空)→受付欄
            date_pat = re.compile(r"(明治|大正|昭和|平成|令和)\d+年")
            ref_pat = re.compile(r"第\d+号")

            if date_pat.search(owner_name) or (not mochibun.startswith("持分") and mochibun and not date_pat.search(mochibun)):
                # 氏名欄に日付がある OR 持分欄に持分以外の値（氏名等）がある → ずれ補正
                # 持分欄→氏名, 氏名欄→日付, 日付欄→受付番号 にシフト
                corrected_name = mochibun
                corrected_date = owner_name
                corrected_ref = reg_date
                corrected_mochibun = ""

                owner_name = corrected_name
                mochibun = corrected_mochibun
                reg_date = corrected_date
                ref_num = corrected_ref

            if owner_name or owner_addr:
                properties[norm_key]["owners"].append({
                    "address": owner_addr,
                    "name": owner_name,
                    "mochibun": mochibun,
                    "reg_date": reg_date,
                    "ref_num": ref_num,
                })

    return properties


def compare_text(old_val: str, new_val: str) -> bool:
    """テキスト比較（正規化後）"""
    return normalize_text(old_val) != normalize_text(new_val)


def get_primary_owner(owners: list) -> tuple:
    """最初の所有者（代表所有者）を取得"""
    if not owners:
        return ("", "")
    return (owners[0].get("name", ""), owners[0].get("address", ""))


def create_excel(source: dict, ocr: dict, output_path: Path):
    """Excel照合レポート生成"""
    wb = Workbook()

    # === 全物件のキーを統合 ===
    all_keys = sorted(set(list(source.keys()) + list(ocr.keys())))

    # 変更一覧データを収集
    changes = []
    full_comparison = []
    unmatched = []

    for key in all_keys:
        src = source.get(key)
        ocr_data = ocr.get(key)

        if src and ocr_data:
            # 両方にある → 比較
            src_chimoku = src.get("chimoku", "")
            ocr_chimoku = ocr_data.get("chimoku", "")
            src_chiseki = src.get("chiseki", "")
            ocr_chiseki = ocr_data.get("chiseki", "")
            src_name, src_addr = get_primary_owner(src.get("owners", []))
            ocr_name, ocr_addr = get_primary_owner(ocr_data.get("owners", []))
            page = ocr_data.get("page", "")

            chimoku_changed = compare_text(src_chimoku, ocr_chimoku) if (src_chimoku and ocr_chimoku) else False
            chiseki_changed = compare_text(src_chiseki, ocr_chiseki) if (src_chiseki and ocr_chiseki) else False
            owner_changed = compare_text(src_name, ocr_name) if (src_name and ocr_name) else False

            # 所在（照合用）
            shozai_for_change = ocr_data.get("shozai", "") or src.get("shozai_full", "") or src.get("shozai", "")

            # 地目のOCR旧値誤読を検出・補正
            ocr_chimoku_corrected = ocr_chimoku
            chimoku_ocr_error = False
            if chimoku_changed and ocr_chimoku:
                src_old_chimoku_values = [normalize_text(v) for v in src.get("chimoku_old_values", [])]
                ocr_chimoku_normalized = normalize_text(ocr_chimoku)
                if ocr_chimoku_normalized in src_old_chimoku_values:
                    # OCRが旧地目（下線付き）を読んでしまった → 照合元の現在地目で補正
                    ocr_chimoku_corrected = src_chimoku
                    chimoku_changed = False
                    chimoku_ocr_error = True

            # 変更一覧に追加
            if chimoku_changed:
                changes.append({
                    "chiban": key, "shozai": shozai_for_change, "item": "地目",
                    "old": src_chimoku, "new": ocr_chimoku, "page": page,
                })
            if chimoku_ocr_error:
                changes.append({
                    "chiban": key, "shozai": shozai_for_change, "item": "地目（OCR補正）",
                    "old": f"OCR読取値: {ocr_chimoku}",
                    "new": f"補正後: {src_chimoku}（照合元と一致）", "page": page,
                })
            if chiseki_changed:
                changes.append({
                    "chiban": key, "shozai": shozai_for_change, "item": "地積",
                    "old": src_chiseki, "new": ocr_chiseki, "page": page,
                })
            if owner_changed:
                changes.append({
                    "chiban": key, "shozai": shozai_for_change, "item": "所有者",
                    "old": src_name, "new": ocr_name, "page": page,
                })

            # 全件照合
            shozai = ocr_data.get("shozai", "") or src.get("shozai_full", "") or src.get("shozai", "")
            full_comparison.append({
                "chiban": key,
                "shozai": shozai,
                "chimoku_old": src_chimoku,
                "chimoku_new": ocr_chimoku_corrected,
                "chimoku_changed": chimoku_changed,
                "chimoku_ocr_error": chimoku_ocr_error,
                "chiseki_old": src_chiseki,
                "chiseki_new": ocr_chiseki,
                "chiseki_changed": chiseki_changed,
                "owner_old": src_name,
                "owner_new": ocr_name,
                "owner_changed": owner_changed,
                "addr_old": src_addr,
                "addr_new": ocr_addr,
                "page": page,
                "source_file": src.get("source_file", ""),
                "owner_count_old": len(src.get("owners", [])),
                "owner_count_new": len(ocr_data.get("owners", [])),
            })

        elif src and not ocr_data:
            unmatched.append({
                "category": "照合元のみ",
                "chiban": key,
                "shozai": src.get("shozai_full", "") or src.get("shozai", ""),
                "source": src.get("source_file", ""),
                "note": "OCR結果に該当なし（PDF範囲外の可能性）",
            })
        elif ocr_data and not src:
            unmatched.append({
                "category": "OCRのみ",
                "chiban": key,
                "shozai": ocr_data.get("shozai", ""),
                "source": f"PDF p.{ocr_data.get('page', '?')}",
                "note": "照合元に該当なし（新規取得・範囲外の可能性）",
            })

    # === シート1: 変更一覧 ===
    ws1 = wb.active
    ws1.title = "変更一覧"
    headers1 = ["所在", "地番", "項目", "照合元（旧）", "OCR結果（新）", "PDFページ", "確認（○/×）", "備考"]
    _write_headers(ws1, headers1)

    for i, ch in enumerate(changes, start=2):
        ws1.cell(row=i, column=1, value=ch.get("shozai", "")).font = NORMAL_FONT
        ws1.cell(row=i, column=2, value=ch["chiban"]).font = NORMAL_FONT
        ws1.cell(row=i, column=3, value=ch["item"]).font = NORMAL_FONT
        ws1.cell(row=i, column=4, value=ch["old"]).font = NORMAL_FONT
        ws1.cell(row=i, column=5, value=ch["new"]).font = NORMAL_FONT
        ws1.cell(row=i, column=6, value=ch["page"]).font = NORMAL_FONT
        ws1.cell(row=i, column=7).font = NORMAL_FONT
        ws1.cell(row=i, column=8).font = NORMAL_FONT

        # ハイライト（OCR補正行は青系、通常変更は黄系）
        if "OCR補正" in str(ch.get("item", "")):
            info_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            ws1.cell(row=i, column=4).fill = info_fill
            ws1.cell(row=i, column=5).fill = info_fill
        else:
            ws1.cell(row=i, column=4).fill = CHANGE_FILL
            ws1.cell(row=i, column=5).fill = CHANGE_FILL

        for col in range(1, 9):
            ws1.cell(row=i, column=col).border = THIN_BORDER

    _auto_width(ws1, headers1)
    ws1.auto_filter.ref = f"A1:H{max(len(changes) + 1, 2)}"

    # === シート2: 全件照合 ===
    ws2 = wb.create_sheet("全件照合")
    headers2 = [
        "地番", "所在", "地目(旧)", "地目(新)", "地目変更",
        "地積(旧)", "地積(新)", "地積変更",
        "所有者(旧)", "所有者(新)", "所有者変更",
        "住所(旧)", "住所(新)",
        "共有者数(旧)", "共有者数(新)",
        "PDFページ", "照合元ファイル",
    ]
    _write_headers(ws2, headers2)

    for i, fc in enumerate(full_comparison, start=2):
        values = [
            fc["chiban"], fc["shozai"],
            fc["chimoku_old"], fc["chimoku_new"], "●" if fc["chimoku_changed"] else "",
            fc["chiseki_old"], fc["chiseki_new"], "●" if fc["chiseki_changed"] else "",
            fc["owner_old"], fc["owner_new"], "●" if fc["owner_changed"] else "",
            fc["addr_old"], fc["addr_new"],
            fc["owner_count_old"], fc["owner_count_new"],
            fc["page"], fc["source_file"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        # 変更セルをハイライト
        if fc["chimoku_changed"]:
            ws2.cell(row=i, column=3).fill = CHANGE_FILL
            ws2.cell(row=i, column=4).fill = CHANGE_FILL
            ws2.cell(row=i, column=5).fill = CHANGE_FILL
        if fc["chiseki_changed"]:
            ws2.cell(row=i, column=6).fill = CHANGE_FILL
            ws2.cell(row=i, column=7).fill = CHANGE_FILL
            ws2.cell(row=i, column=8).fill = CHANGE_FILL
        if fc["owner_changed"]:
            ws2.cell(row=i, column=9).fill = CHANGE_FILL
            ws2.cell(row=i, column=10).fill = CHANGE_FILL
            ws2.cell(row=i, column=11).fill = CHANGE_FILL

    _auto_width(ws2, headers2)
    ws2.auto_filter.ref = f"A1:Q{max(len(full_comparison) + 1, 2)}"

    # === シート3: 照合不能 ===
    ws3 = wb.create_sheet("照合不能")
    headers3 = ["区分", "地番", "所在", "出典", "備考"]
    _write_headers(ws3, headers3)

    for i, um in enumerate(unmatched, start=2):
        values = [um["category"], um["chiban"], um["shozai"], um["source"], um["note"]]
        for col, val in enumerate(values, start=1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        fill = NEW_FILL if um["category"] == "OCRのみ" else MISSING_FILL
        for col in range(1, 6):
            ws3.cell(row=i, column=col).fill = fill

    _auto_width(ws3, headers3)
    ws3.auto_filter.ref = f"A1:E{max(len(unmatched) + 1, 2)}"

    # 保存
    wb.save(output_path)

    # サマリー出力
    print(f"\n{'='*60}")
    print(f"照合レポート生成完了")
    print(f"{'='*60}")
    print(f"  出力先: {output_path}")
    print(f"\n  シート1「変更一覧」: {len(changes)}件の変更検出")
    print(f"  シート2「全件照合」: {len(full_comparison)}物件を横並び比較")
    print(f"  シート3「照合不能」: {len(unmatched)}件")

    # 照合不能の内訳
    ocr_only = sum(1 for u in unmatched if u["category"] == "OCRのみ")
    src_only = sum(1 for u in unmatched if u["category"] == "照合元のみ")
    print(f"    - OCRのみ（照合元になし）: {ocr_only}件")
    print(f"    - 照合元のみ（OCRになし）: {src_only}件")

    print(f"\n  照合元物件数: {len(source)}")
    print(f"  OCR物件数: {len(ocr)}")
    print(f"  マッチ数: {len(full_comparison)}")


def _write_headers(ws, headers):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_width(ws, headers):
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(h) * 2  # 日本語は2倍幅
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) * 1.5)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def main():
    print("照合元CSV読み込み中...")
    source = parse_source_csvs(SOURCE_CSV_DIR)
    print(f"  {len(source)}物件をパース")

    print("OCR結果CSV読み込み中...")
    ocr = parse_ocr_csv(OCR_CSV)
    print(f"  {len(ocr)}物件をパース")

    print("\n照合・比較中...")
    create_excel(source, ocr, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
