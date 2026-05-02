"""
登記事項要約書 照合レポート生成 v2
照合元CSV（調査開始時点）× OCR結果CSV（最新 v2）→ Excel 3シート

v1からの改善点:
- 氏名比較: スペース除去で突合（OCR側・CSV側とも）
- OCR CSV: 9列固定フォーマット（持分列の位置ずれ解消）
- 受付情報: 日付+番号を1列に統合
- 地目OCR旧値誤読の自動補正を継続
- 外字コード変換を継続
"""

import csv
import re
import unicodedata
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# === 設定 ===
SOURCE_CSV_DIR = Path("/home/maaatan/Chiseki/youyakusyo/csv")
OCR_CSV = Path("/home/maaatan/Chiseki/youyakusyo/output_v2/要約書_全データ.csv")
OUTPUT_XLSX = Path("/home/maaatan/Chiseki/youyakusyo/output_v2/照合レポート.xlsx")

# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
CHANGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NEW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
OCR_FIX_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def zen_to_han(text: str) -> str:
    """全角数字・英字・記号を半角に変換"""
    return unicodedata.normalize("NFKC", text)


def normalize_chiban(raw: str) -> str:
    """地番を正規化"""
    s = zen_to_han(raw).strip()
    s = s.replace("－", "-").replace("ー", "-").replace("—", "-")
    s = re.sub(r"(\d+)-(\d+)", r"\1番\2", s)
    return s


def chiban_match_key(raw: str) -> str:
    """地番をマッチング用に正規化"""
    s = normalize_chiban(raw)
    s = s.replace("番の", "番")
    s = re.sub(r"番$", "", s)
    return s


def normalize_empty_marker(text: str) -> str:
    """空欄マーカー `(空)` を空文字列に変換"""
    v = (text or "").strip()
    if v in ("(空)", "（空）", "(空欄)", "（空欄）", "-", "―"):
        return ""
    return v


def normalize_text(text: str) -> str:
    """テキスト正規化（比較用）"""
    s = normalize_empty_marker(text)
    s = zen_to_han(s).strip()
    s = s.replace("　", " ").replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("北\u8328城", "北茨城")
    return s


# 住所判定用キーワード
ADDRESS_KEYWORDS = re.compile(
    r"(都|道|府|県|市|区|町|村|郡|番地|丁目|番の|番地の|字|大字)"
)


def looks_like_address(text: str) -> bool:
    """住所っぽい文字列か判定"""
    if not text:
        return False
    if len(text) < 4:
        return False
    return bool(ADDRESS_KEYWORDS.search(text))


def looks_like_name(text: str) -> bool:
    """氏名っぽい文字列か判定（住所キーワードを含まず、短い）"""
    if not text:
        return False
    if looks_like_address(text):
        return False
    if len(text) > 30:  # 長すぎるものは住所等
        return False
    return True


def normalize_chiseki(text: str) -> str:
    """地積正規化: 中黒→ピリオド、全角→半角"""
    s = zen_to_han(text).strip()
    s = s.replace("・", ".").replace("‧", ".")
    # 小数点以下の不要な0を除去
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def normalize_name(text: str) -> str:
    """氏名正規化: 全角半角統一 + スペース完全除去"""
    s = normalize_text(text)
    s = s.replace(" ", "")
    return s


# 異体字マッピング（旧字体/互換漢字 → 新字体/標準漢字）
VARIANT_MAP = {
    "﨑": "崎", "髙": "高", "藏": "蔵", "惠": "恵", "鐵": "鉄",
    "廣": "広", "寬": "寛", "眞": "真", "實": "実", "彌": "弥",
    "經": "経", "濱": "浜", "澤": "沢", "邊": "辺", "齋": "斎",
    "齊": "斉", "國": "国", "嶋": "島", "條": "条", "學": "学",
    "萬": "万", "壽": "寿", "榮": "栄", "禮": "礼", "顯": "顕",
    "黑": "黒", "靜": "静", "關": "関", "シヨ": "ショ",
}


def is_variant_match(text_a: str, text_b: str) -> bool:
    """異体字を正規化した場合に一致するかを判定（自動補正はしない）"""
    a = normalize_name(text_a)
    b = normalize_name(text_b)
    if a == b:
        return False  # そもそも一致しているので異体字判定不要
    for old, new in VARIANT_MAP.items():
        a = a.replace(old, new)
        b = b.replace(old, new)
    return a == b


# 外字コードマッピング
GAIJI_MAP = {
    "<00008328>": "茨",
    "<00009AD9>": "髙",
    "<0000FA11>": "﨑",
    "<0000FA19>": "神",
    "<0000F9DC>": "隆",
    "<00005BEC>": "寬",
    "<00005EE3>": "廣",
    "<000080D6>": "胖",
    "<00008FC4>": "迄",
    "<000049FA>": "䧺",
    "<0000E452>": "〓",
    "<00100F20>": "〓",
    "<000F61BE>": "〓",
}


def replace_gaiji(text: str) -> str:
    """外字コード置換"""
    for code, char in GAIJI_MAP.items():
        text = text.replace(code, char)

    def _replace_unknown(m):
        try:
            cp = int(m.group(1), 16)
            if 0 < cp < 0x110000:
                return chr(cp)
        except (ValueError, OverflowError):
            pass
        return "〓"

    text = re.sub(r"<([0-9A-Fa-f]{8})>", _replace_unknown, text)
    return text


def parse_source_csvs(csv_dir: Path) -> dict:
    """照合元CSV（Shift-JIS階層構造）をパース"""
    properties = {}

    for csv_path in sorted(csv_dir.glob("*_001_BM.csv")):
        with open(csv_path, "r", encoding="cp932", errors="replace") as f:
            reader = csv.reader(f)
            current_prop = None
            current_num = None

            for row in reader:
                if len(row) < 3:
                    continue

                row = [replace_gaiji(cell) for cell in row]
                num = zen_to_han(row[0]).strip()
                record_type = zen_to_han(row[1]).strip()

                if record_type == "物件情報":
                    # 前の物件を保存
                    if current_prop:
                        _save_property(properties, current_prop)

                    raw_chiban = zen_to_han(row[5]).strip() if len(row) > 5 else ""
                    shozai = normalize_text(row[4]) if len(row) > 4 else ""
                    aza = normalize_text(row[9]) if len(row) > 9 else ""

                    current_num = num
                    current_prop = {
                        "source_file": csv_path.name,
                        "raw_chiban": raw_chiban,
                        "shozai": shozai,
                        "aza": aza,
                        "chimoku": "",
                        "chiseki": "",
                        "chimoku_old_values": [],
                        "owners": [],
                        "history": [],
                    }

                elif record_type.startswith("所在") and current_prop:
                    full_shozai = normalize_text(row[2]) if len(row) > 2 else ""
                    if full_shozai:
                        current_prop["shozai_full"] = full_shozai

                elif record_type.startswith("表示履歴") and current_prop:
                    chiban_val = zen_to_han(row[2]).strip() if len(row) > 2 else ""
                    chimoku_raw = zen_to_han(row[4]).strip() if len(row) > 4 else ""
                    chiseki_raw = zen_to_han(row[5]).strip() if len(row) > 5 else ""

                    if chiban_val and not current_prop.get("chiban_display"):
                        current_prop["chiban_display"] = chiban_val

                    is_old_chimoku = bool(re.search(r"[【】]", chimoku_raw))
                    is_old_chiseki = bool(re.search(r"[【】]", chiseki_raw))
                    chimoku_clean = re.sub(r"[【】\[\]]", "", chimoku_raw).strip()
                    chiseki_clean = re.sub(r"[【】\[\]]", "", chiseki_raw).strip()

                    if chimoku_clean and not is_old_chimoku:
                        current_prop["chimoku"] = chimoku_clean
                    elif chimoku_clean and is_old_chimoku:
                        current_prop["chimoku_old_values"].append(chimoku_clean)
                        if not current_prop["chimoku"]:
                            current_prop["chimoku_old"] = chimoku_clean

                    if chiseki_clean and not is_old_chiseki:
                        current_prop["chiseki"] = chiseki_clean
                    elif chiseki_clean and is_old_chiseki and not current_prop["chiseki"]:
                        current_prop["chiseki_old"] = chiseki_clean

                    current_prop["history"].append({
                        "chimoku": chimoku_raw, "chiseki": chiseki_raw,
                        "detail": zen_to_han(row[6]).strip() if len(row) > 6 else "",
                    })

                elif record_type.startswith("所有権") and current_prop:
                    addr = normalize_text(row[2]) if len(row) > 2 else ""
                    mochibun = normalize_text(row[3]) if len(row) > 3 else ""
                    name = normalize_text(row[4]) if len(row) > 4 else ""
                    reg_info = normalize_text(row[6]) if len(row) > 6 else ""
                    ref_num = normalize_text(row[7]) if len(row) > 7 else ""

                    if name or addr:
                        current_prop["owners"].append({
                            "address": addr,
                            "mochibun": mochibun,
                            "name": name,
                            "reg_info": reg_info,
                            "ref_num": ref_num,
                        })

                elif record_type.startswith("所有者") and current_prop:
                    name = normalize_text(row[2]) if len(row) > 2 else ""
                    if name:
                        current_prop["owners"].append({
                            "address": "",
                            "mochibun": "",
                            "name": name,
                            "reg_info": "",
                            "ref_num": "",
                        })

            # 最後の物件
            if current_prop:
                _save_property(properties, current_prop)

    return properties


def _save_property(properties: dict, prop: dict):
    """物件を辞書に保存"""
    chiban_key = prop.get("chiban_display", prop["raw_chiban"])
    if chiban_key:
        norm_key = chiban_match_key(chiban_key)
        properties[norm_key] = prop


def detect_number_gaps(csv_path: Path) -> list:
    """対策5: 番号欠落検知
    OCR結果CSVを走査し、ページ内・PDFファイル内で番号が飛んでいる箇所を検出
    """
    gaps = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        current_pdf = None
        prev_num = 0
        prev_page = ""
        prev_chiban = ""
        seen_nums = set()

        for row in reader:
            pdf_file = (row.get("PDFファイル") or "").strip()
            page = (row.get("ページ") or "").strip()
            num_raw = (row.get("番号") or "").strip()
            chiban = (row.get("地番") or "").strip()

            # PDFファイルが変わったらリセット
            if pdf_file != current_pdf:
                current_pdf = pdf_file
                prev_num = 0
                seen_nums = set()

            # 数字のみ
            if not num_raw.isdigit():
                continue
            num = int(num_raw)

            # 同じ番号の繰り返し（共有者の複数行）はスキップ
            if num in seen_nums:
                continue
            seen_nums.add(num)

            # 番号の飛びを検出
            if prev_num > 0 and num > prev_num + 1:
                gap_size = num - prev_num - 1
                missing_nums = list(range(prev_num + 1, num))
                # 判定: 大きな欠落（>5）は整理番号セクション境界の可能性大
                # 小さな欠落（1-5）はOCR読み飛ばしの可能性大
                if gap_size <= 5:
                    severity = "OCR読み飛ばしの可能性 — 要PDF目視確認"
                else:
                    severity = f"大きな欠落 — 整理番号（セクション）境界の可能性。目視で整理番号の変化を確認"
                gaps.append({
                    "pdf_file": pdf_file,
                    "page_before": prev_page,
                    "page_after": page,
                    "prev_num": prev_num,
                    "prev_chiban": prev_chiban,
                    "next_num": num,
                    "next_chiban": chiban,
                    "missing_nums": missing_nums,
                    "gap_size": gap_size,
                    "severity": severity,
                })

            prev_num = num
            prev_page = page
            prev_chiban = chiban

    return gaps


def parse_ocr_csv(csv_path: Path) -> dict:
    """OCR結果CSV（v2: 9列固定）をパース"""
    properties = {}

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 空欄マーカー正規化
            chiban_raw = normalize_empty_marker(row.get("地番") or "")
            if not chiban_raw:
                continue

            # 不正な地番フィルタ
            chiban_check = zen_to_han(chiban_raw)
            if not re.match(r"^\d", chiban_check):
                continue
            if "番地" in chiban_raw:
                continue

            norm_key = chiban_match_key(chiban_raw)

            if norm_key not in properties:
                properties[norm_key] = {
                    "shozai": normalize_empty_marker(row.get("所在") or ""),
                    "chiban_display": chiban_raw,
                    "chimoku": normalize_empty_marker(row.get("地目") or ""),
                    "chiseki": normalize_empty_marker(row.get("地積") or ""),
                    "owners": [],
                    "page": (row.get("ページ") or ""),
                    "pdf_file": (row.get("PDFファイル") or ""),
                }

            # 所有者追加（カラムずれ補正付き）
            owner_addr = normalize_empty_marker(row.get("住所") or "")
            mochibun = normalize_empty_marker(row.get("持分") or "")
            owner_name = normalize_empty_marker(row.get("氏名") or "")
            uketsuke = normalize_empty_marker(row.get("受付情報") or "")

            # バグ1-B対策: 住所列に氏名が入るケース（問題A: 1083番馬上初吉）
            # パターン: 住所=氏名っぽい短文, 持分・氏名・受付すべて空
            # → 住所列の値を氏名列に移動
            if (owner_addr and
                not mochibun and
                not owner_name and
                not uketsuke and
                not looks_like_address(owner_addr) and
                looks_like_name(owner_addr)):
                owner_name = owner_addr
                owner_addr = ""

            # カラムずれ検出・補正
            date_pat = re.compile(r"(明治|大正|昭和|平成|令和)\d+年")
            ref_pat = re.compile(r"第\d+号")

            # バグ1対策: 持分列に氏名が入るケース（パターンA・C）
            # パターンA: 住所=空, 持分=氏名, 氏名=空
            # パターンC: 住所=あり, 持分=氏名, 氏名=空
            # → いずれも持分列の値を氏名列に移動
            if (mochibun and
                not mochibun.startswith("持分") and
                not date_pat.search(mochibun) and
                not ref_pat.search(mochibun) and
                not owner_name):
                owner_name = mochibun
                mochibun = ""

            if date_pat.search(owner_name) or ref_pat.search(owner_name):
                # 氏名列に受付情報が入っている → ずれ補正
                if mochibun and not mochibun.startswith("持分") and not date_pat.search(mochibun):
                    # 持分列に氏名が入っている: 住所,氏名(→持分),受付(→氏名)
                    uketsuke = owner_name
                    owner_name = mochibun
                    mochibun = ""
                else:
                    # 住所列に氏名、持分=空、氏名列に受付: 氏名(→住所),,受付(→氏名)
                    if owner_addr and not date_pat.search(owner_addr):
                        uketsuke = owner_name
                        owner_name = owner_addr
                        owner_addr = ""
                        mochibun = ""
                    else:
                        # 補正できない → 受付情報として扱い、氏名は空
                        uketsuke = owner_name
                        owner_name = ""

            if owner_name or owner_addr:
                properties[norm_key]["owners"].append({
                    "address": owner_addr,
                    "mochibun": mochibun,
                    "name": owner_name,
                    "uketsuke": uketsuke,
                })

    return properties


def compare_text(old_val: str, new_val: str) -> bool:
    """テキスト比較（正規化後）"""
    return normalize_text(old_val) != normalize_text(new_val)


def compare_name(old_val: str, new_val: str) -> bool:
    """氏名比較（スペース除去して比較）"""
    return normalize_name(old_val) != normalize_name(new_val)


def chiseki_cross_reference(src_chiseki: str, ocr_chiseki: str, chimoku: str) -> str:
    """対策5: 地積の照合元クロスリファレンス補正
    宅地/雑種地でOCR値が照合元の小数点読み落とし・連結の場合、照合元値を返す
    """
    src_n = normalize_chiseki(src_chiseki)
    ocr_n = normalize_chiseki(ocr_chiseki)

    if src_n == ocr_n:
        return ""  # 既に一致

    if "." not in src_n:
        return ""  # 照合元が整数なら補正対象外

    # 小数点付き地目チェック（宅地・雑種地が主）
    # ただし他の地目でも小数点がある場合があるので、地目に関わらずチェックする

    src_int = src_n.split(".")[0]
    src_dec = src_n.split(".")[1] if "." in src_n else ""

    # パターン1: 切り捨て型（OCRが整数部のみ）
    if ocr_n == src_int:
        return src_chiseki  # 照合元値を採用

    # パターン2: 連結型（OCRが整数部+小数部を連結）
    concatenated = src_int + src_dec
    if ocr_n == concatenated:
        return src_chiseki  # 照合元値を採用

    return ""  # 補正不可


def compare_owners(src_owners: list, ocr_owners: list) -> list:
    """対策3: 全所有者の差分比較
    返り値: 変更リスト [{type: "変更"|"追加"|"消失", ...}]
    """
    diffs = []
    src_names = {normalize_name(o.get("name", "")): o for o in src_owners if o.get("name")}
    ocr_names = {normalize_name(o.get("name", "")): o for o in ocr_owners if o.get("name")}

    src_set = set(src_names.keys())
    ocr_set = set(ocr_names.keys())

    # 一致する所有者（変更なし）
    matched = src_set & ocr_set

    # 照合元のみ（消失）
    for name_key in sorted(src_set - ocr_set):
        src_o = src_names[name_key]
        # 異体字チェック: OCR側に異体字で一致するものがないか
        variant_found = False
        for ocr_key in sorted(ocr_set - src_set):
            ocr_o = ocr_names[ocr_key]
            if is_variant_match(src_o.get("name", ""), ocr_o.get("name", "")):
                variant_found = True
                diffs.append({
                    "type": "異体字差異",
                    "src_name": src_o.get("name", ""),
                    "ocr_name": ocr_o.get("name", ""),
                    "src_addr": src_o.get("address", ""),
                    "ocr_addr": ocr_o.get("address", ""),
                })
                ocr_set.discard(ocr_key)
                break
        if not variant_found:
            diffs.append({
                "type": "消失",
                "src_name": src_o.get("name", ""),
                "ocr_name": "",
                "src_addr": src_o.get("address", ""),
                "ocr_addr": "",
            })

    # OCRのみ（追加）
    for name_key in sorted(ocr_set - src_set - matched):
        ocr_o = ocr_names[name_key]
        diffs.append({
            "type": "追加",
            "src_name": "",
            "ocr_name": ocr_o.get("name", ""),
            "src_addr": "",
            "ocr_addr": ocr_o.get("address", ""),
        })

    return diffs


def create_excel(source: dict, ocr: dict, gaps: list, output_path: Path):
    """Excel照合レポート生成"""
    wb = Workbook()
    all_keys = sorted(set(list(source.keys()) + list(ocr.keys())))

    changes = []
    full_comparison = []
    unmatched = []

    for key in all_keys:
        src = source.get(key)
        ocr_data = ocr.get(key)

        if src and ocr_data:
            src_chimoku = src.get("chimoku", "")
            ocr_chimoku = ocr_data.get("chimoku", "")
            src_chiseki = src.get("chiseki", "")
            ocr_chiseki = ocr_data.get("chiseki", "")
            page = ocr_data.get("page", "")
            shozai = ocr_data.get("shozai", "") or src.get("shozai_full", "") or src.get("shozai", "")

            # --- 地目比較 ---
            chimoku_changed = compare_text(src_chimoku, ocr_chimoku) if (src_chimoku and ocr_chimoku) else False
            ocr_chimoku_corrected = ocr_chimoku
            chimoku_ocr_error = False
            if chimoku_changed and ocr_chimoku:
                src_old_values = [normalize_text(v) for v in src.get("chimoku_old_values", [])]
                if normalize_text(ocr_chimoku) in src_old_values:
                    ocr_chimoku_corrected = src_chimoku
                    chimoku_changed = False
                    chimoku_ocr_error = True

            # --- 地積比較（対策5: クロスリファレンス補正）---
            chiseki_changed = (
                normalize_chiseki(src_chiseki) != normalize_chiseki(ocr_chiseki)
                if (src_chiseki and ocr_chiseki) else False
            )
            chiseki_corrected = ""
            if chiseki_changed:
                chiseki_corrected = chiseki_cross_reference(
                    src_chiseki, ocr_chiseki, ocr_chimoku or src_chimoku
                )
                if chiseki_corrected:
                    chiseki_changed = False  # 小数点読み落としと判定、変更から除外

            # --- 所有者比較（対策3: 全員差分）---
            src_owners = src.get("owners", [])
            ocr_owners = ocr_data.get("owners", [])
            owner_diffs = compare_owners(src_owners, ocr_owners)

            # 変更一覧に追加
            if chimoku_changed:
                changes.append({
                    "chiban": key, "shozai": shozai, "item": "地目",
                    "old": src_chimoku, "new": ocr_chimoku, "page": page,
                    "note": "",
                })
            if chimoku_ocr_error:
                changes.append({
                    "chiban": key, "shozai": shozai, "item": "地目（OCR補正）",
                    "old": f"OCR読取値: {ocr_chimoku}",
                    "new": f"補正後: {src_chimoku}（照合元と一致）", "page": page,
                    "note": "",
                })
            if chiseki_changed:
                changes.append({
                    "chiban": key, "shozai": shozai, "item": "地積",
                    "old": src_chiseki, "new": ocr_chiseki, "page": page,
                    "note": "",
                })
            if chiseki_corrected:
                changes.append({
                    "chiban": key, "shozai": shozai, "item": "地積（小数点補正）",
                    "old": f"照合元: {src_chiseki}",
                    "new": f"OCR読取値: {ocr_chiseki}（整数部一致、小数点読み落とし）",
                    "page": page,
                    "note": "小数点の縦点線がOCRで認識されなかった可能性",
                })

            # 対策3+4: 所有者差分を全件記録
            for diff in owner_diffs:
                if diff["type"] == "異体字差異":
                    changes.append({
                        "chiban": key, "shozai": shozai, "item": "所有者（異体字）",
                        "old": diff["src_name"],
                        "new": diff["ocr_name"],
                        "page": page,
                        "note": "異体字差異の可能性あり — 要目視確認",
                    })
                elif diff["type"] == "消失":
                    changes.append({
                        "chiban": key, "shozai": shozai, "item": "所有者（消失）",
                        "old": diff["src_name"],
                        "new": "（OCRに該当なし）",
                        "page": page,
                        "note": f"照合元住所: {diff['src_addr']}" if diff["src_addr"] else "",
                    })
                elif diff["type"] == "追加":
                    changes.append({
                        "chiban": key, "shozai": shozai, "item": "所有者（追加）",
                        "old": "（照合元に該当なし）",
                        "new": diff["ocr_name"],
                        "page": page,
                        "note": f"OCR住所: {diff['ocr_addr']}" if diff["ocr_addr"] else "",
                    })

            # 全件照合用データ
            src_name_list = ", ".join(o.get("name", "") for o in src_owners if o.get("name"))
            ocr_name_list = ", ".join(o.get("name", "") for o in ocr_owners if o.get("name"))
            has_owner_diff = len(owner_diffs) > 0

            full_comparison.append({
                "chiban": key,
                "shozai": shozai,
                "chimoku_old": src_chimoku,
                "chimoku_new": ocr_chimoku_corrected,
                "chimoku_changed": chimoku_changed,
                "chimoku_ocr_error": chimoku_ocr_error,
                "chiseki_old": src_chiseki,
                "chiseki_new": ocr_chiseki if not chiseki_corrected else f"{ocr_chiseki}→補正:{chiseki_corrected}",
                "chiseki_changed": chiseki_changed,
                "owner_old": src_name_list,
                "owner_new": ocr_name_list,
                "owner_changed": has_owner_diff,
                "page": page,
                "source_file": src.get("source_file", ""),
                "owner_count_old": len(src_owners),
                "owner_count_new": len(ocr_owners),
            })

        elif src and not ocr_data:
            unmatched.append({
                "category": "照合元のみ",
                "chiban": key,
                "shozai": src.get("shozai_full", "") or src.get("shozai", ""),
                "source": src.get("source_file", ""),
                "note": "OCR結果に該当なし（PDF範囲外の可能性）",
            })
        elif ocr_data and not src:
            unmatched.append({
                "category": "OCRのみ",
                "chiban": key,
                "shozai": ocr_data.get("shozai", ""),
                "source": f"PDF p.{ocr_data.get('page', '?')}",
                "note": "照合元に該当なし（新規取得・範囲外の可能性）",
            })

    # === シート1: 変更一覧 ===
    ws1 = wb.active
    ws1.title = "変更一覧"
    headers1 = ["所在", "地番", "項目", "照合元（旧）", "OCR結果（新）", "PDFページ", "確認（○/×）", "備考"]
    _write_headers(ws1, headers1)

    VARIANT_FILL = PatternFill(start_color="E8D5E8", end_color="E8D5E8", fill_type="solid")
    CHISEKI_FIX_FILL = PatternFill(start_color="D5E8D5", end_color="D5E8D5", fill_type="solid")

    for i, ch in enumerate(changes, start=2):
        vals = [ch.get("shozai", ""), ch["chiban"], ch["item"],
                ch["old"], ch["new"], ch["page"], "",
                ch.get("note", "")]
        for col, val in enumerate(vals, start=1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        item = str(ch.get("item", ""))
        if "OCR補正" in item:
            ws1.cell(row=i, column=4).fill = OCR_FIX_FILL
            ws1.cell(row=i, column=5).fill = OCR_FIX_FILL
        elif "異体字" in item:
            ws1.cell(row=i, column=4).fill = VARIANT_FILL
            ws1.cell(row=i, column=5).fill = VARIANT_FILL
        elif "小数点補正" in item:
            ws1.cell(row=i, column=4).fill = CHISEKI_FIX_FILL
            ws1.cell(row=i, column=5).fill = CHISEKI_FIX_FILL
        else:
            ws1.cell(row=i, column=4).fill = CHANGE_FILL
            ws1.cell(row=i, column=5).fill = CHANGE_FILL

    _auto_width(ws1, headers1)
    ws1.auto_filter.ref = f"A1:H{max(len(changes) + 1, 2)}"

    # === シート2: 全件照合 ===
    ws2 = wb.create_sheet("全件照合")
    headers2 = [
        "地番", "所在", "地目(旧)", "地目(新)", "地目変更",
        "地積(旧)", "地積(新)", "地積変更",
        "所有者一覧(旧)", "所有者一覧(新)", "所有者変更",
        "人数(旧)", "人数(新)",
        "PDFページ", "照合元ファイル",
    ]
    _write_headers(ws2, headers2)

    for i, fc in enumerate(full_comparison, start=2):
        values = [
            fc["chiban"], fc["shozai"],
            fc["chimoku_old"], fc["chimoku_new"], "●" if fc["chimoku_changed"] else "",
            fc["chiseki_old"], fc["chiseki_new"], "●" if fc["chiseki_changed"] else "",
            fc["owner_old"], fc["owner_new"], "●" if fc["owner_changed"] else "",
            fc["owner_count_old"], fc["owner_count_new"],
            fc["page"], fc["source_file"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        if fc["chimoku_changed"]:
            for c in [3, 4, 5]:
                ws2.cell(row=i, column=c).fill = CHANGE_FILL
        if fc["chiseki_changed"]:
            for c in [6, 7, 8]:
                ws2.cell(row=i, column=c).fill = CHANGE_FILL
        if fc["owner_changed"]:
            for c in [9, 10, 11]:
                ws2.cell(row=i, column=c).fill = CHANGE_FILL

    _auto_width(ws2, headers2)
    ws2.auto_filter.ref = f"A1:O{max(len(full_comparison) + 1, 2)}"

    # === シート3: 番号欠落（対策5）===
    ws_gap = wb.create_sheet("番号欠落")
    headers_gap = [
        "PDFファイル", "前ページ", "次ページ", "前の番号", "前の地番",
        "次の番号", "次の地番", "欠落番号", "備考",
    ]
    _write_headers(ws_gap, headers_gap)

    GAP_SMALL_FILL = PatternFill(start_color="FFB6B6", end_color="FFB6B6", fill_type="solid")
    GAP_LARGE_FILL = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
    for i, g in enumerate(gaps, start=2):
        vals = [
            g["pdf_file"], g["page_before"], g["page_after"],
            g["prev_num"], g["prev_chiban"],
            g["next_num"], g["next_chiban"],
            ", ".join(str(n) for n in g["missing_nums"]),
            f"欠落{g['gap_size']}件 — {g['severity']}",
        ]
        fill = GAP_SMALL_FILL if g["gap_size"] <= 5 else GAP_LARGE_FILL
        for col, val in enumerate(vals, start=1):
            cell = ws_gap.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.fill = fill

    _auto_width(ws_gap, headers_gap)
    ws_gap.auto_filter.ref = f"A1:I{max(len(gaps) + 1, 2)}"

    # === シート4: 照合不能 ===
    ws3 = wb.create_sheet("照合不能")
    headers3 = ["区分", "地番", "所在", "出典", "備考"]
    _write_headers(ws3, headers3)

    for i, um in enumerate(unmatched, start=2):
        vals = [um["category"], um["chiban"], um["shozai"], um["source"], um["note"]]
        for col, val in enumerate(vals, start=1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        fill = NEW_FILL if um["category"] == "OCRのみ" else MISSING_FILL
        for col in range(1, 6):
            ws3.cell(row=i, column=col).fill = fill

    _auto_width(ws3, headers3)
    ws3.auto_filter.ref = f"A1:E{max(len(unmatched) + 1, 2)}"

    # 保存
    wb.save(output_path)

    # サマリー
    print(f"\n{'='*60}")
    print(f"照合レポート生成完了")
    print(f"{'='*60}")
    print(f"  出力先: {output_path}")
    print(f"\n  シート1「変更一覧」: {len(changes)}件の変更検出")

    # 変更内訳
    chimoku_real = sum(1 for c in changes if c["item"] == "地目")
    chimoku_fix = sum(1 for c in changes if c["item"] == "地目（OCR補正）")
    chiseki_changes = sum(1 for c in changes if c["item"] == "地積")
    owner_changes = sum(1 for c in changes if c["item"] == "所有者")
    print(f"    地目変更: {chimoku_real}件, 地目OCR補正: {chimoku_fix}件")
    print(f"    地積変更: {chiseki_changes}件, 所有者変更: {owner_changes}件")

    # 変更内訳
    chimoku_real = sum(1 for c in changes if c["item"] == "地目")
    chimoku_fix = sum(1 for c in changes if c["item"] == "地目（OCR補正）")
    chiseki_real = sum(1 for c in changes if c["item"] == "地積")
    chiseki_fix = sum(1 for c in changes if c["item"] == "地積（小数点補正）")
    owner_variant = sum(1 for c in changes if c["item"] == "所有者（異体字）")
    owner_lost = sum(1 for c in changes if c["item"] == "所有者（消失）")
    owner_added = sum(1 for c in changes if c["item"] == "所有者（追加）")
    print(f"    地目変更: {chimoku_real}件, 地目OCR補正: {chimoku_fix}件")
    print(f"    地積変更: {chiseki_real}件, 地積小数点補正: {chiseki_fix}件")
    print(f"    所有者 異体字（要目視）: {owner_variant}件, 消失: {owner_lost}件, 追加: {owner_added}件")

    print(f"  シート2「全件照合」: {len(full_comparison)}物件")
    print(f"  シート3「照合不能」: {len(unmatched)}件")

    ocr_only = sum(1 for u in unmatched if u["category"] == "OCRのみ")
    src_only = sum(1 for u in unmatched if u["category"] == "照合元のみ")
    print(f"    OCRのみ: {ocr_only}件, 照合元のみ: {src_only}件")

    print(f"\n  照合元物件数: {len(source)}")
    print(f"  OCR物件数: {len(ocr)}")
    print(f"  マッチ数: {len(full_comparison)}")


def _write_headers(ws, headers):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_width(ws, headers):
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(h) * 2
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) * 1.5)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def main():
    print("照合元CSV読み込み中...")
    source = parse_source_csvs(SOURCE_CSV_DIR)
    print(f"  {len(source)}物件をパース")

    print("OCR結果CSV読み込み中...")
    ocr = parse_ocr_csv(OCR_CSV)
    print(f"  {len(ocr)}物件をパース")

    print("番号欠落検知中...")
    gaps = detect_number_gaps(OCR_CSV)
    print(f"  {len(gaps)}件の番号欠落を検出")

    print("\n照合・比較中...")
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    create_excel(source, ocr, gaps, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
