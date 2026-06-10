#!/usr/bin/env python3
"""
Access MDB差分ツールのブラウザUI。

Windowsでは pyodbc と Access Database Engine でMDBを読み取る。
Linux検証環境では mdbtools があれば同じ処理を実行できる。
"""

from __future__ import annotations

import base64
import csv
import io
import json
import shutil
import subprocess
import tempfile
import traceback
import webbrowser
from dataclasses import dataclass
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    print("openpyxl が見つかりません。初回セットアップ.bat を実行してください。")
    raise


HOST = "127.0.0.1"
PORT = 8767
DEFAULT_TABLES = "L31,L33,L34"
DEFAULT_IGNORE_COLUMNS = "ID"
LAND_KEY_CANDIDATES = [
    ["AzaKY", "FDKBN", "CHIBN", "SEQ", "NUMB"],
    ["AzaKY", "FDKBN", "CHIBN", "SEQ", "Numb"],
    ["AzaKY", "FDKBN", "CHIBN", "SEQ"],
    ["mAzaKY", "MFDKBN", "MCHIBN", "MSEQ"],
    ["mAzaKY", "MCHIBN"],
    ["CHIBN", "SEQ", "NUMB"],
    ["CHIBN", "SEQ"],
    ["MCHIBN", "MSEQ"],
    ["USRCD"],
]

# 全テーブル共通の「筆」識別キー。L31を参照源として、L33/L34等の差分にも
# 地番・地目・地積・字を付与するために使う。
LAND_PARCEL_KEY_COLS = ["AzaKY", "FDKBN", "CHIBN", "SEQ"]
LAND_PARCEL_KEY_COLS_ALT = ["mAzaKY", "MFDKBN", "MCHIBN", "MSEQ"]
LAND_CONTEXT_FIELDS = ("地番", "地目", "地積", "字")


HTML = """<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>ACCESS差分ツール</title>
  <style>
    body {
      margin: 0;
      font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #f4f6f8;
      color: #1f2937;
    }
    main {
      max-width: 980px;
      margin: 28px auto;
      padding: 0 18px;
    }
    h1 {
      font-size: 24px;
      margin: 0 0 18px;
      letter-spacing: 0;
    }
    section {
      background: #fff;
      border: 1px solid #d7dde5;
      border-radius: 8px;
      padding: 18px;
      margin-bottom: 14px;
    }
    .grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }
    label {
      display: block;
      font-weight: 650;
      margin-bottom: 8px;
    }
    input[type=file], input[type=text] {
      width: 100%;
      box-sizing: border-box;
      padding: 10px;
      border: 1px solid #c7ced8;
      border-radius: 6px;
      background: #fff;
      font-size: 15px;
    }
    button {
      border: 0;
      border-radius: 6px;
      background: #215d49;
      color: #fff;
      padding: 11px 18px;
      font-size: 15px;
      font-weight: 700;
      cursor: pointer;
    }
    button:disabled {
      background: #9aa8bd;
      cursor: wait;
    }
    pre {
      min-height: 170px;
      overflow: auto;
      white-space: pre-wrap;
      word-break: break-word;
      background: #111827;
      color: #e5e7eb;
      border-radius: 6px;
      padding: 14px;
      font-size: 14px;
      line-height: 1.55;
    }
    .hint {
      margin-top: 8px;
      color: #4b5563;
      font-size: 13px;
    }
    @media (max-width: 760px) {
      main { margin-top: 18px; }
      .grid { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <main>
    <h1>ACCESS差分ツール</h1>
    <section class="grid">
      <div>
        <label for="baseFile">基準MDB（例: DT2.MDB・最新側）</label>
        <input id="baseFile" type="file" accept=".mdb,.accdb,.MDB,.ACCDB">
      </div>
      <div>
        <label for="compareFile">比較MDB（例: マスター・チェックリスト側）</label>
        <input id="compareFile" type="file" accept=".mdb,.accdb,.MDB,.ACCDB">
      </div>
    </section>
    <section class="grid">
      <div>
        <label for="tables">比較対象テーブル（カンマ区切り）</label>
        <input id="tables" type="text" value="L31,L33,L34">
        <div class="hint">空欄にすると、両方に存在する同名テーブルをすべて比較します。</div>
      </div>
      <div>
        <label for="ignoreColumns">無視する列（カンマ区切り）</label>
        <input id="ignoreColumns" type="text" value="ID">
        <div class="hint">片方だけにある管理用IDなどを除外します。</div>
      </div>
    </section>
    <section>
      <button id="runButton">差分Excelを作成</button>
    </section>
    <section>
      <label>結果</label>
      <pre id="log">MDBファイルを2つ選択して「差分Excelを作成」を押してください。</pre>
    </section>
  </main>
  <script>
    const baseFile = document.getElementById("baseFile");
    const compareFile = document.getElementById("compareFile");
    const tables = document.getElementById("tables");
    const ignoreColumns = document.getElementById("ignoreColumns");
    const runButton = document.getElementById("runButton");
    const log = document.getElementById("log");

    function readAsBase64(file) {
      return new Promise((resolve, reject) => {
        const reader = new FileReader();
        reader.onload = () => resolve(String(reader.result).split(",", 2)[1]);
        reader.onerror = reject;
        reader.readAsDataURL(file);
      });
    }

    function downloadBase64(filename, base64Data) {
      const binary = atob(base64Data);
      const bytes = new Uint8Array(binary.length);
      for (let i = 0; i < binary.length; i++) {
        bytes[i] = binary.charCodeAt(i);
      }
      const blob = new Blob([bytes], {
        type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
      });
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = filename;
      document.body.appendChild(a);
      a.click();
      a.remove();
      URL.revokeObjectURL(url);
    }

    runButton.addEventListener("click", async () => {
      if (!baseFile.files[0] || !compareFile.files[0]) {
        log.textContent = "MDBファイルを2つ選択してください。";
        return;
      }

      runButton.disabled = true;
      log.textContent = "MDBを読み込み、差分を作成しています。ファイルが大きい場合は数分かかります...";

      try {
        const payload = {
          base_name: baseFile.files[0].name,
          base_data: await readAsBase64(baseFile.files[0]),
          compare_name: compareFile.files[0].name,
          compare_data: await readAsBase64(compareFile.files[0]),
          tables: tables.value,
          ignore_columns: ignoreColumns.value
        };
        const response = await fetch("/compare", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload)
        });
        const result = await response.json();
        if (!response.ok) {
          throw new Error(result.error || "差分作成に失敗しました。");
        }
        log.textContent = result.summary;
        downloadBase64(result.output_name, result.output_data);
      } catch (error) {
        log.textContent = "エラーが発生しました。\\n\\n" + error.message;
      } finally {
        runButton.disabled = false;
      }
    });
  </script>
</body>
</html>
"""


@dataclass
class TableData:
    columns: list[str]
    rows: list[dict[str, Any]]


@dataclass
class TableDiff:
    table: str
    status: str
    key_columns: list[str]
    base_count: int = 0
    compare_count: int = 0
    changed_count: int = 0
    base_only_count: int = 0
    compare_only_count: int = 0
    diffs: list[dict[str, Any]] | None = None
    base_only: list[dict[str, Any]] | None = None
    compare_only: list[dict[str, Any]] | None = None
    message: str = ""


def split_csv_text(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.hex()
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def key_to_text(key: tuple[str, ...]) -> str:
    return " / ".join(key)


def _first_nonempty(row: dict[str, Any], candidates: list[str]) -> str:
    for col in candidates:
        if col in row:
            val = normalize_cell(row.get(col))
            if val:
                return val
    return ""


def _decode_chibn(chibn: str) -> str:
    """CHIBN(15桁コード)を本番-枝番形式に整形する。
    例: '000373000000000' -> '373', '001220000010000' -> '1220-10'
    形式が想定外なら入力をそのまま返す。"""
    if not chibn or len(chibn) != 15:
        return chibn
    head = chibn[0:6].lstrip("0") or "0"
    mid = chibn[6:12].lstrip("0")
    tail = chibn[12:15].lstrip("0")
    parts = [head]
    if mid:
        parts.append(mid)
    if tail:
        parts.append(tail)
    return "-".join(parts)


def extract_land_attrs(row: dict[str, Any]) -> dict[str, str]:
    """L31等の1行から地番・地目・地積・字を取り出す。
    ChibNM/MOKUNMが空の場合はコード値(CHIBN/CHICD)へフォールバックする。"""
    chibnm = _first_nonempty(row, ["ChibNM", "mChibNM"])
    chibn_raw = _first_nonempty(row, ["CHIBN", "MCHIBN"])
    chiban = chibnm or _decode_chibn(chibn_raw) or chibn_raw

    mokunm = _first_nonempty(row, ["MOKUNM"])
    chicd = _first_nonempty(row, ["CHICD", "MCHICD"])
    if mokunm:
        chimoku = mokunm
    elif chicd and chicd != "0":
        chimoku = f"コード:{chicd}"
    else:
        chimoku = ""

    return {
        "地番": chiban,
        "地目": chimoku,
        "地積": _first_nonempty(row, ["TOKSK", "KEISK", "MCHISK"]),
        "字": _first_nonempty(row, ["a小字名", "m小字名"]),
    }


def row_parcel_key(row: dict[str, Any]) -> tuple[str, ...] | None:
    """任意行から(AzaKY, FDKBN, CHIBN, SEQ)形式の筆キーを推定する。"""
    for cols in (LAND_PARCEL_KEY_COLS, LAND_PARCEL_KEY_COLS_ALT):
        if all(col in row for col in cols):
            key = tuple(normalize_cell(row.get(col)) for col in cols)
            if any(key):
                return key
    return None


def build_land_context_index(
    base_l31: "TableData | None",
    compare_l31: "TableData | None",
) -> dict[tuple[str, ...], dict[str, str]]:
    """L31から筆キー→(地番/地目/地積/字)の辞書を作る。基準MDB側を優先。"""
    index: dict[tuple[str, ...], dict[str, str]] = {}
    # compare_l31 を先に入れて base_l31 で上書き（base優先）
    for source in (compare_l31, base_l31):
        if source is None:
            continue
        if not all(col in source.columns for col in LAND_PARCEL_KEY_COLS):
            continue
        for row in source.rows:
            key = row_parcel_key(row)
            if key is None:
                continue
            attrs = extract_land_attrs(row)
            if not any(attrs.values()) and key in index:
                continue
            index[key] = attrs
    return index


def lookup_land_context(
    row: dict[str, Any] | None,
    index: dict[tuple[str, ...], dict[str, str]],
) -> dict[str, str]:
    empty = {field: "" for field in LAND_CONTEXT_FIELDS}
    if not row:
        return empty
    # まずは行自身に地番/地目/地積/字が入っているか（L31本体の場合）を確認
    direct = extract_land_attrs(row)
    key = row_parcel_key(row)
    fallback = index.get(key, empty) if key is not None else empty
    return {field: direct.get(field) or fallback.get(field, "") for field in LAND_CONTEXT_FIELDS}


class AccessReader:
    def __init__(self, path: Path):
        self.path = path
        self._conn = None
        self._mode = "odbc"
        if shutil.which("mdb-tables") and shutil.which("mdb-export"):
            self._mode = "mdbtools"

    def __enter__(self) -> "AccessReader":
        if self._mode == "odbc":
            try:
                import pyodbc
            except ModuleNotFoundError as exc:
                raise RuntimeError("pyodbc が見つかりません。初回セットアップ.bat を実行してください。") from exc
            drivers = [d for d in pyodbc.drivers() if "Access Driver" in d and ("*.mdb" in d or "*.accdb" in d)]
            if not drivers:
                raise RuntimeError("Access ODBCドライバが見つかりません。Microsoft Access Database Engine 64bit をインストールしてください。")
            driver = drivers[-1]
            conn_str = f"DRIVER={{{driver}}};DBQ={self.path};"
            self._conn = pyodbc.connect(conn_str, autocommit=True)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if self._conn is not None:
            self._conn.close()

    def table_names(self) -> list[str]:
        if self._mode == "mdbtools":
            result = subprocess.run(["mdb-tables", "-1", str(self.path)], text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            if result.returncode != 0:
                raise RuntimeError(result.stderr.strip() or "mdb-tables に失敗しました。")
            return [line.strip() for line in result.stdout.splitlines() if line.strip()]
        cursor = self._conn.cursor()
        return sorted(row.table_name for row in cursor.tables(tableType="TABLE") if not row.table_name.startswith("MSys"))

    def read_table(self, table: str) -> TableData:
        if self._mode == "mdbtools":
            result = subprocess.run(["mdb-export", str(self.path), table], text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            if result.returncode != 0:
                raise RuntimeError(result.stderr.strip() or f"{table} の読み込みに失敗しました。")
            reader = csv.DictReader(io.StringIO(result.stdout))
            columns = list(reader.fieldnames or [])
            rows = [{col: normalize_cell(row.get(col)) for col in columns} for row in reader]
            return TableData(columns=columns, rows=rows)
        cursor = self._conn.cursor()
        sql = f"SELECT * FROM [{table}]"
        cursor.execute(sql)
        columns = [col[0] for col in cursor.description]
        rows = []
        for row in cursor.fetchall():
            rows.append({col: normalize_cell(value) for col, value in zip(columns, row)})
        return TableData(columns=columns, rows=rows)


def choose_key_columns(columns: list[str], rows: list[dict[str, Any]], ignore: set[str]) -> list[str]:
    usable = [col for col in columns if col not in ignore]
    for candidate in LAND_KEY_CANDIDATES:
        if all(col in usable for col in candidate):
            keys = [tuple(normalize_cell(row.get(col)) for col in candidate) for row in rows]
            non_empty = [key for key in keys if any(key)]
            if non_empty and len(non_empty) == len(set(non_empty)):
                return candidate
    for candidate in [["ID"], ["id"]]:
        if all(col in columns for col in candidate) and not any(col in ignore for col in candidate):
            keys = [tuple(normalize_cell(row.get(col)) for col in candidate) for row in rows]
            if keys and len(keys) == len(set(keys)):
                return candidate
    usable_sorted = sorted(usable)
    if not usable_sorted:
        return columns[:1]
    return usable_sorted


def build_index(rows: list[dict[str, Any]], key_columns: list[str]) -> dict[tuple[str, ...], dict[str, Any]]:
    index: dict[tuple[str, ...], dict[str, Any]] = {}
    seen: dict[tuple[str, ...], int] = {}
    for row in rows:
        key = tuple(normalize_cell(row.get(col)) for col in key_columns)
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            key = key + (f"重複{seen[key]}",)
        index[key] = row
    return index


def compare_table(table: str, base: TableData, compare: TableData, ignore: set[str]) -> TableDiff:
    base_columns = [col for col in base.columns if col not in ignore]
    compare_columns = [col for col in compare.columns if col not in ignore]
    common_columns = [col for col in base_columns if col in set(compare_columns)]
    key_columns = choose_key_columns(common_columns, base.rows or compare.rows, ignore)
    compare_columns_set = set(compare_columns)

    base_index = build_index(base.rows, key_columns)
    compare_index = build_index(compare.rows, key_columns)
    base_keys = set(base_index)
    compare_keys = set(compare_index)
    shared_keys = sorted(base_keys & compare_keys)

    diffs: list[dict[str, Any]] = []
    changed_keys: set[tuple[str, ...]] = set()
    for key in shared_keys:
        base_row = base_index[key]
        compare_row = compare_index[key]
        for col in common_columns:
            if col in key_columns:
                continue
            base_value = normalize_cell(base_row.get(col))
            compare_value = normalize_cell(compare_row.get(col))
            if base_value != compare_value:
                changed_keys.add(key)
                diffs.append({
                    "table": table,
                    "key": key_to_text(key),
                    "column": col,
                    "base_value": base_value,
                    "compare_value": compare_value,
                    "base_row": base_row,
                    "compare_row": compare_row,
                })

    base_only = [base_index[key] for key in sorted(base_keys - compare_keys)]
    compare_only = [compare_index[key] for key in sorted(compare_keys - base_keys)]
    missing_columns = [col for col in base_columns if col not in compare_columns_set]
    extra_columns = [col for col in compare_columns if col not in set(base_columns)]
    message_parts = []
    if missing_columns:
        message_parts.append("比較MDBにない列: " + ", ".join(missing_columns))
    if extra_columns:
        message_parts.append("比較MDBだけの列: " + ", ".join(extra_columns))

    return TableDiff(
        table=table,
        status="比較完了",
        key_columns=key_columns,
        base_count=len(base.rows),
        compare_count=len(compare.rows),
        changed_count=len(changed_keys),
        base_only_count=len(base_only),
        compare_only_count=len(compare_only),
        diffs=diffs,
        base_only=base_only,
        compare_only=compare_only,
        message=" / ".join(message_parts),
    )


def compare_databases(
    base_path: Path,
    compare_path: Path,
    requested_tables: list[str],
    ignore_columns: list[str],
) -> tuple[list[TableDiff], dict[tuple[str, ...], dict[str, str]], list[str], list[str]]:
    ignore = set(ignore_columns)
    with AccessReader(base_path) as base_reader, AccessReader(compare_path) as compare_reader:
        base_tables = base_reader.table_names()
        compare_tables = compare_reader.table_names()
        base_set = set(base_tables)
        compare_set = set(compare_tables)

        # 先にL31を読み、筆キー→属性のコンテキスト索引を構築する。
        # 他テーブル（L33/L34等）の差分行にも地番/地目/地積/字を付加するため。
        l31_base: TableData | None = None
        l31_compare: TableData | None = None
        if "L31" in base_set:
            try:
                l31_base = base_reader.read_table("L31")
            except Exception:
                l31_base = None
        if "L31" in compare_set:
            try:
                l31_compare = compare_reader.read_table("L31")
            except Exception:
                l31_compare = None
        context_index = build_land_context_index(l31_base, l31_compare)

        tables = requested_tables or sorted(base_set & compare_set)
        results: list[TableDiff] = []
        for table in tables:
            if table not in base_set:
                results.append(TableDiff(table=table, status="基準MDBにテーブルなし", key_columns=[], message="比較をスキップしました。"))
                continue
            if table not in compare_set:
                results.append(TableDiff(table=table, status="比較MDBにテーブルなし", key_columns=[], message="比較をスキップしました。"))
                continue
            # L31は既に読み込み済みなら再利用する。
            if table == "L31" and l31_base is not None:
                base_data = l31_base
            else:
                base_data = base_reader.read_table(table)
            if table == "L31" and l31_compare is not None:
                compare_data = l31_compare
            else:
                compare_data = compare_reader.read_table(table)
            results.append(compare_table(table, base_data, compare_data, ignore))
        return results, context_index, base_tables, compare_tables


def append_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="DDE7F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows[:200]:
            if idx <= len(row):
                max_len = max(max_len, len(str(row[idx - 1])) if row[idx - 1] is not None else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 55)


def rows_for_table_only(
    results: list[TableDiff],
    side: str,
    context_index: dict[tuple[str, ...], dict[str, str]],
) -> list[list[Any]]:
    out: list[list[Any]] = []
    for result in results:
        rows = result.base_only if side == "base" else result.compare_only
        for row in rows or []:
            ctx = lookup_land_context(row, context_index)
            out.append([
                result.table,
                ctx["地番"], ctx["地目"], ctx["地積"], ctx["字"],
                json.dumps(row, ensure_ascii=False),
            ])
    return out


def build_workbook(
    results: list[TableDiff],
    context_index: dict[tuple[str, ...], dict[str, str]],
    base_tables: list[str],
    compare_tables: list[str],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = []
    for result in results:
        summary_rows.append([
            result.table,
            result.status,
            ", ".join(result.key_columns),
            result.base_count,
            result.compare_count,
            result.changed_count,
            result.base_only_count,
            result.compare_only_count,
            result.message,
        ])
    append_sheet(wb, "概要", ["テーブル", "状態", "照合キー", "基準件数", "比較件数", "変更行数", "基準のみ", "比較のみ", "備考"], summary_rows)

    diff_headers = ["テーブル", "照合キー"] + list(LAND_CONTEXT_FIELDS) + ["列", "基準MDBの値", "比較MDBの値"]
    diff_rows = []
    for result in results:
        for diff in result.diffs or []:
            ctx = lookup_land_context(diff.get("base_row") or diff.get("compare_row"), context_index)
            diff_rows.append([
                diff["table"], diff["key"],
                ctx["地番"], ctx["地目"], ctx["地積"], ctx["字"],
                diff["column"], diff["base_value"], diff["compare_value"],
            ])
    append_sheet(wb, "差分", diff_headers, diff_rows)

    only_headers = ["テーブル"] + list(LAND_CONTEXT_FIELDS) + ["行データ(JSON)"]
    append_sheet(wb, "基準のみ", only_headers, rows_for_table_only(results, "base", context_index))
    append_sheet(wb, "比較のみ", only_headers, rows_for_table_only(results, "compare", context_index))

    table_rows = []
    all_tables = sorted(set(base_tables) | set(compare_tables))
    for table in all_tables:
        table_rows.append([table, "あり" if table in base_tables else "", "あり" if table in compare_tables else ""])
    append_sheet(wb, "テーブル一覧", ["テーブル", "基準MDB", "比較MDB"], table_rows)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def make_summary(results: list[TableDiff]) -> str:
    lines = ["差分Excelを作成しました。", ""]
    for result in results:
        if result.status != "比較完了":
            lines.append(f"{result.table}: {result.status}")
            continue
        lines.append(
            f"{result.table}: 基準{result.base_count}件 / 比較{result.compare_count}件 / "
            f"変更{result.changed_count}行 / 基準のみ{result.base_only_count}件 / 比較のみ{result.compare_only_count}件"
        )
        if result.message:
            lines.append(f"  {result.message}")
    return "\n".join(lines)


def run_compare(payload: dict[str, Any]) -> dict[str, Any]:
    tables = split_csv_text(payload.get("tables") or "")
    ignore_columns = split_csv_text(payload.get("ignore_columns") or "")
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        base_name = payload.get("base_name") or "base.mdb"
        compare_name = payload.get("compare_name") or "compare.mdb"
        base_path = temp_path / base_name
        compare_path = temp_path / compare_name
        base_path.write_bytes(base64.b64decode(payload["base_data"]))
        compare_path.write_bytes(base64.b64decode(payload["compare_data"]))

        results, context_index, base_tables, compare_tables = compare_databases(base_path, compare_path, tables, ignore_columns)
        workbook = build_workbook(results, context_index, base_tables, compare_tables)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return {
            "summary": make_summary(results),
            "output_name": f"ACCESS差分_{stamp}.xlsx",
            "output_data": base64.b64encode(workbook).decode("ascii"),
        }


class Handler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        if self.path != "/":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        body = HTML.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self) -> None:
        if self.path != "/compare":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            result = run_compare(payload)
            self._send_json(HTTPStatus.OK, result)
        except Exception as exc:
            traceback.print_exc()
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})

    def _send_json(self, status: HTTPStatus, payload: dict[str, Any]) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args: Any) -> None:
        print("%s - %s" % (self.address_string(), format % args))


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    url = f"http://{HOST}:{PORT}/"
    print(f"ACCESS差分ツールを起動しました: {url}")
    webbrowser.open(url)
    server.serve_forever()


if __name__ == "__main__":
    main()
