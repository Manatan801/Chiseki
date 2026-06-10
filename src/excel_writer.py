"""Excel出力モジュール

結線指示票（.xls）と交点計算指示書（.xlsx）のExcel出力を行う。
テンプレートファイルをコピーし、計算結果データを書き込む。
"""

from __future__ import annotations

import os
import re
import unicodedata
from copy import copy
from dataclasses import dataclass

import xlrd
import xlwt
from xlutils.copy import copy as xlutils_copy
from xlutils.filter import process, XLRDReader, XLWTWriter

import openpyxl


# ============================================================
# データクラス定義
# ============================================================

@dataclass
class KessenHeaderInfo:
    """結線指示票のヘッダー共通情報"""
    year: str | None = None           # 年度（例: "令和５年度"）
    district: str | None = None       # 地区名（例: "小豆畑・上小津田地区"）
    block_number: float | str | None = None  # ブロック番号
    oaza_name: str | None = None      # 大字名（例: "華川町小豆畑"）
    aza_name: str | None = None       # 字名（例: "内城"）


@dataclass
class KessenResult:
    """結線指示票の1筆分データ"""
    parcel_number: str          # 地番
    stake_sequence: list[str]   # 杭番号の順序リスト（閉合）
    land_use: str | None = None # 地目
    owner: str | None = None    # 所有者名


@dataclass
class KoutenHeaderInfo:
    """交点計算指示書のヘッダー共通情報"""
    district: str | None = None       # 地区名（例: "小豆畑・上小津田"）
    recorder: str | None = None       # 記入者名（例: "前田　高行"）
    block_number: str | None = None   # ブロック番号
    oaza_name: str | None = None      # 大字名（例: "華川町小豆畑"）


@dataclass
class IntersectionResult:
    """交点計算指示書の1交点分データ"""
    intersection_stake: str     # 交点杭番号
    baseline_point1: str        # 基準線の点1
    baseline_point2: str        # 基準線の点2
    extension_point1: str       # 延長線の点1
    extension_point2: str       # 延長線の点2


# ============================================================
# 結線指示票 (.xls) 出力
# ============================================================

def _get_style_list(rb: xlrd.Book) -> list[xlwt.XFStyle]:
    """xlrdブックからxlwt用のスタイルリストを取得する。

    xlutils.filterのprocessを使い、xlrdのXF情報を
    xlwtのXFStyleオブジェクトに変換する。
    """
    reader = XLRDReader(rb, "unknown.xls")
    writer = XLWTWriter()
    process(reader, writer)
    return writer.style_list


def _set_xlwt_attr(obj, attr: str, value) -> None:
    """xlwt objects expose some settings as setters and some as attributes."""
    setter = getattr(obj, f"set_{attr}", None)
    try:
        if setter is not None:
            setter(value)
        elif hasattr(obj, attr):
            setattr(obj, attr, value)
    except Exception:
        # Not every BIFF setting exposed by xlrd can be written by xlwt.
        pass


def _copy_xls_sheet_layout(
    rs: xlrd.sheet.Sheet,
    ws: xlwt.Worksheet,
    style_list: list[xlwt.XFStyle],
) -> None:
    """テンプレートシートの行高・列幅・表示設定をコピーする。"""
    for col_idx, colinfo in rs.colinfo_map.items():
        col = ws.col(col_idx)
        col.width = colinfo.width
        col.hidden = bool(colinfo.hidden)
        col.level = colinfo.outline_level
        col.collapse = bool(colinfo.collapsed)
        if 0 <= colinfo.xf_index < len(style_list):
            col.set_style(style_list[colinfo.xf_index])

    for row_idx, rowinfo in rs.rowinfo_map.items():
        row = ws.row(row_idx)
        row.height = rowinfo.height
        row.height_mismatch = bool(rowinfo.height_mismatch)
        row.hidden = bool(rowinfo.hidden)
        row.level = rowinfo.outline_level
        row.collapse = bool(rowinfo.outline_group_starts_ends)
        row.has_default_height = bool(rowinfo.has_default_height)
        row.space_above = bool(rowinfo.additional_space_above)
        row.space_below = bool(rowinfo.additional_space_below)
        if (
            rowinfo.has_default_xf_index
            and 0 <= rowinfo.xf_index < len(style_list)
        ):
            row.set_style(style_list[rowinfo.xf_index])

    default_mappings = [
        ("default_row_height", "row_default_height"),
        ("default_row_height_mismatch", "row_default_height_mismatch"),
        ("default_row_hidden", "row_default_hidden"),
        ("default_additional_space_above", "row_default_space_above"),
        ("default_additional_space_below", "row_default_space_below"),
        ("defcolwidth", "col_default_width"),
    ]
    for source_attr, target_attr in default_mappings:
        if hasattr(rs, source_attr):
            _set_xlwt_attr(ws, target_attr, getattr(rs, source_attr))

    sheet_mappings = [
        ("visibility", "sheet_visible"),
        ("show_formulas", "show_formulas"),
        ("show_grid_lines", "show_grid"),
        ("show_sheet_headers", "show_headers"),
        ("panes_are_frozen", "panes_frozen"),
        ("remove_splits_if_pane_freeze_is_removed", "remove_splits"),
        ("horz_split_pos", "horz_split_pos"),
        ("vert_split_pos", "vert_split_pos"),
        ("horz_split_first_visible", "horz_split_first_visible"),
        ("vert_split_first_visible", "vert_split_first_visible"),
        ("first_visible_rowx", "first_visible_row"),
        ("first_visible_colx", "first_visible_col"),
        ("automatic_grid_line_colour", "auto_colour_grid"),
        ("gridline_colour_index", "grid_colour"),
        ("columns_from_right_to_left", "cols_right_to_left"),
        ("show_outline_symbols", "show_outline"),
        ("cached_normal_view_mag_factor", "normal_magn"),
        ("cached_page_break_preview_mag_factor", "preview_magn"),
        ("scl_mag_factor", "scl_magn"),
        ("show_in_page_break_preview", "page_preview"),
        ("sheet_selected", "selected"),
        ("horizontal_page_breaks", "horz_page_breaks"),
        ("vertical_page_breaks", "vert_page_breaks"),
    ]
    for source_attr, target_attr in sheet_mappings:
        if hasattr(rs, source_attr):
            _set_xlwt_attr(ws, target_attr, getattr(rs, source_attr))


def _copy_template_sheet(
    rb: xlrd.Book,
    wb: xlwt.Workbook,
    style_list: list[xlwt.XFStyle],
    sheet_name: str,
) -> xlwt.Worksheet:
    """テンプレートの最初のシートを書式付きで新シートにコピーする。

    Args:
        rb: テンプレートのxlrdブック
        wb: 出力先のxlwtワークブック
        style_list: XFStyleリスト
        sheet_name: 新シートの名前

    Returns:
        書式コピー済みの新しいxlwtワークシート
    """
    rs = rb.sheet_by_index(0)
    ws = wb.add_sheet(sheet_name, cell_overwrite_ok=True)
    _copy_xls_sheet_layout(rs, ws, style_list)

    # 全セルの値とスタイルをコピー
    for row_idx in range(rs.nrows):
        for col_idx in range(rs.ncols):
            xf_idx = rs.cell_xf_index(row_idx, col_idx)
            val = rs.cell_value(row_idx, col_idx)
            if xf_idx < len(style_list):
                ws.write(row_idx, col_idx, val, style_list[xf_idx])
            else:
                ws.write(row_idx, col_idx, val)

    # セル結合をコピー
    for rlo, rhi, clo, chi in rs.merged_cells:
        ws.merge(rlo, rhi - 1, clo, chi - 1)

    return ws


def _parcel_sort_tuple(parcel_number: str) -> tuple[int, ...] | None:
    """地番を自然順ソート用の数値タプルに変換する。"""
    text = unicodedata.normalize("NFKC", str(parcel_number).strip())
    if not text or not text[0].isdigit():
        return None
    if not re.fullmatch(r"\d+(?:[-番の]+\d+)*", text):
        return None
    return tuple(int(part) for part in re.findall(r"\d+", text))


def _sort_kessen_results(results: list[KessenResult]) -> list[KessenResult]:
    """読み取れた地番を若い順にし、不明地番は末尾で元順を保つ。"""
    def sort_key(item: tuple[int, KessenResult]):
        index, result = item
        parcel_key = _parcel_sort_tuple(result.parcel_number)
        if parcel_key is None:
            return (1, (), index)
        return (0, parcel_key, index)

    return [result for _, result in sorted(enumerate(results), key=sort_key)]


def _write_kessen_sheet(
    ws: xlwt.Worksheet,
    result: KessenResult,
    style_list: list[xlwt.XFStyle],
    rb: xlrd.Book,
    header: KessenHeaderInfo | None = None,
    page_offset: int = 0,
) -> None:
    """結線指示票の1シートにデータを書き込む。

    杭番号データは行5~29（0-indexed）の4列配置:
      Col A(0): 番号1-25, Col B(1): 杭番号
      Col D(3): 番号26-50, Col E(4): 杭番号
      Col G(6): 番号51-75, Col H(7): 杭番号
      Col J(9): 番号76-100, Col K(10): 杭番号

    page_offset: 杭番号リストの開始位置（100点超の2ページ目は100）
    """
    rs = rb.sheet_by_index(0)

    def _get_style(row: int, col: int) -> xlwt.XFStyle:
        xf_idx = rs.cell_xf_index(row, col)
        if xf_idx < len(style_list):
            return style_list[xf_idx]
        return xlwt.XFStyle()

    # ヘッダー情報の書き込み
    if header is not None:
        if header.year is not None:
            ws.write(1, 0, header.year, _get_style(1, 0))
        if header.district is not None:
            ws.write(1, 3, header.district, _get_style(1, 3))
        if header.block_number is not None:
            ws.write(3, 0, header.block_number, _get_style(3, 0))
        if header.oaza_name is not None:
            ws.write(3, 2, header.oaza_name, _get_style(3, 2))
        if header.aza_name is not None:
            ws.write(3, 4, header.aza_name, _get_style(3, 4))

    # ブロック番号を自動判定してA4に記入（headerに未指定の場合）
    if (header is None or header.block_number is None) and result.stake_sequence:
        # 杭番号の最頻ブロック番号を採用
        from collections import Counter
        block_counts = Counter()
        for sn in result.stake_sequence:
            b = _extract_block(sn)
            if b is not None:
                block_counts[b] += 1
        if block_counts:
            auto_block = block_counts.most_common(1)[0][0]
            ws.write(3, 0, auto_block, _get_style(3, 0))

    # 地番をRow 3, Col 6に書き込み
    ws.write(3, 6, result.parcel_number, _get_style(3, 6))

    # 杭番号データの書き込み
    # 4列配置: 各列25行（行5~29、0-indexed）
    column_configs = [
        (1,),    # 列B: 杭番号 1-25
        (4,),    # 列E: 杭番号 26-50
        (7,),    # 列H: 杭番号 51-75
        (10,),   # 列K: 杭番号 76-100
    ]

    stakes = result.stake_sequence[page_offset:page_offset + 100]
    for i, stake in enumerate(stakes):
        col_group = i // 25           # 0, 1, 2, 3
        row_offset = i % 25           # 0-24
        row = 5 + row_offset          # 行5~29（0-indexed）
        stake_col = column_configs[col_group][0]

        # 杭番号セルのスタイルを取得し、文字列書式で書き込み
        # （"1.830"等の末尾ゼロが消えないよう、書式を"@"=テキストに設定）
        xf_idx = rs.cell_xf_index(row, stake_col)
        style = style_list[xf_idx] if xf_idx < len(style_list) else xlwt.XFStyle()
        text_style = xlwt.XFStyle()
        text_style.font = style.font
        text_style.alignment = style.alignment
        text_style.borders = style.borders
        text_style.pattern = style.pattern
        text_style.protection = style.protection
        text_style.num_format_str = '@'
        ws.write(row, stake_col, _strip_stake_prefix(stake), text_style)


def write_kessen_excel(
    results: list[KessenResult],
    template_path: str,
    output_path: str,
    header: KessenHeaderInfo | None = None,
) -> None:
    """結線指示票をExcelに出力する。

    テンプレート(.xls)をコピーし、各筆ごとにシートを作成して
    杭番号データを書き込む。

    Args:
        results: 結線指示票データのリスト
        template_path: テンプレートファイルパス
        output_path: 出力ファイルパス
        header: ヘッダー共通情報（年度、地区名等）
    """
    if not results:
        raise ValueError("結果データが空です")

    # 出力ディレクトリを作成
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # テンプレートを読み込み
    rb = xlrd.open_workbook(template_path, formatting_info=True)
    style_list = _get_style_list(rb)
    sorted_results = _sort_kessen_results(results)

    # xlutils.copyでワークブックをコピー（テンプレートシートの書式保持）
    wb = xlutils_copy(rb)

    def _unique_sheet_name(name: str) -> str:
        """シート名の重複を連番で回避"""
        if name not in used_names:
            used_names[name] = 1
            return name
        used_names[name] += 1
        return f"{name}({used_names[name]})"

    def _write_result_sheets(result: KessenResult, is_first: bool) -> None:
        """1筆分のシートを作成（100点超なら複数ページ）"""
        total = len(result.stake_sequence)
        pages = (total + 99) // 100  # 切り上げ

        for page in range(pages):
            offset = page * 100
            if page == 0 and is_first:
                ws = wb.get_sheet(0)
                ws.name = _unique_sheet_name(result.parcel_number)
            else:
                suffix = f"_{page + 1}" if pages > 1 and page > 0 else ""
                sheet_name = _unique_sheet_name(
                    f"{result.parcel_number}{suffix}"
                )
                ws = _copy_template_sheet(rb, wb, style_list, sheet_name)
            _write_kessen_sheet(
                ws, result, style_list, rb, header, page_offset=offset
            )

    used_names: dict[str, int] = {}
    _write_result_sheets(sorted_results[0], is_first=True)
    for result in sorted_results[1:]:
        _write_result_sheets(result, is_first=False)

    # 保存
    wb.save(output_path)


# ============================================================
# 交点計算指示書 (.xlsx) 出力
# ============================================================

def _to_number(value: str) -> float | str:
    """文字列を可能であれば数値に変換する。

    Excelセルに数値として書き込むため、数値に変換可能な
    文字列はfloatに変換する。変換不可の場合は文字列のまま返す。
    """
    try:
        return float(value)
    except (ValueError, TypeError):
        return value


# 交点ブロック配置の定数
# 横3ブロック x 縦5段 = 最大15交点
_KOUTEN_ROWS_PER_BLOCK = 11     # 段の行間隔
_KOUTEN_COLS_PER_BLOCK = 18     # ブロックの列間隔
_KOUTEN_MAX_COLS = 3            # 横ブロック数
_KOUTEN_MAX_ROWS = 5            # 縦段数
_KOUTEN_MAX_POINTS = _KOUTEN_MAX_COLS * _KOUTEN_MAX_ROWS  # 15

# 各段の中央◎行（1-indexed for openpyxl）
_KOUTEN_CENTER_ROWS = [10, 21, 32, 43, 54]


def _get_kouten_cell_positions(
    block_idx: int,
) -> dict[str, tuple[int, int]]:
    """交点ブロックのデータ書き込み位置を取得する。

    Args:
        block_idx: ブロックインデックス（0~14）

    Returns:
        フィールド名 → (row, col) の辞書（1-indexed）
    """
    row_group = block_idx // _KOUTEN_MAX_COLS  # 段 (0-4)
    col_group = block_idx % _KOUTEN_MAX_COLS   # 列 (0-2)

    center_row = _KOUTEN_CENTER_ROWS[row_group]
    col_offset = col_group * _KOUTEN_COLS_PER_BLOCK

    return {
        # 延長線（方向線）の点1: 上○（◎行-5, 列offset+11）
        "extension_point1": (center_row - 5, col_offset + 11),
        # 延長線（方向線）の点2: 上から2番目の○（◎行-2, 列offset+11）
        "extension_point2": (center_row - 2, col_offset + 11),
        # 基準線の左: 左○（◎行+2, 列offset+2）
        "baseline_point1": (center_row + 2, col_offset + 2),
        # 交点杭番号: ◎（◎行+2, 列offset+8）
        "intersection_stake": (center_row + 2, col_offset + 8),
        # 基準線の右: 右○（◎行+2, 列offset+14）
        "baseline_point2": (center_row + 2, col_offset + 14),
    }


def _strip_stake_prefix(number: str) -> str:
    """杭番号をそのまま返す（接頭辞は除去しない）"""
    return number


def _extract_block(stake_number: str) -> str | None:
    """杭番号からブロック番号を抽出する"""
    cleaned = re.sub(r'^[-交既復計]', '', stake_number)
    m = re.match(r'^(\d+)\.', cleaned)
    return m.group(1) if m else None


def _write_kouten_sheet(
    ws,
    results: list[IntersectionResult],
    block_number: str | None = None,
    header: KoutenHeaderInfo | None = None,
) -> None:
    """交点計算指示書の1シートにデータを書き込む。

    Args:
        ws: openpyxlワークシート
        results: 交点計算結果のリスト（最大15件）
        block_number: ブロック番号（A3セルに記入）
        header: ヘッダー共通情報
    """
    # ヘッダー情報の書き込み（1-indexed）
    if header is not None:
        if header.district is not None:
            ws.cell(row=2, column=14, value=header.district)
        if header.recorder is not None:
            ws.cell(row=2, column=44, value=header.recorder)
        if header.oaza_name is not None:
            ws.cell(row=3, column=22, value=header.oaza_name)

    # ブロック番号をA3に記入（引数 > header の優先順）
    effective_block = block_number
    if effective_block is None and header is not None:
        effective_block = header.block_number
    if effective_block is not None:
        ws.cell(row=3, column=1, value=_to_number(str(effective_block)))

    # 各交点データを書き込み
    for idx, result in enumerate(results):
        if idx >= _KOUTEN_MAX_POINTS:
            break
        positions = _get_kouten_cell_positions(idx)

        for field, value in [
            ("extension_point1", _strip_stake_prefix(result.extension_point1)),
            ("extension_point2", _strip_stake_prefix(result.extension_point2)),
            ("baseline_point1", _strip_stake_prefix(result.baseline_point1)),
            ("intersection_stake", _strip_stake_prefix(result.intersection_stake)),
            ("baseline_point2", _strip_stake_prefix(result.baseline_point2)),
        ]:
            cell = ws.cell(
                row=positions[field][0],
                column=positions[field][1],
                value=value if value else value,
            )
            cell.number_format = '@'


def _copy_openpyxl_attr(source_ws, target_ws, attr: str) -> None:
    try:
        setattr(target_ws, attr, copy(getattr(source_ws, attr)))
    except Exception:
        pass


def _copy_openpyxl_sheet_layout(source_ws, target_ws) -> None:
    """copy_worksheetで落ちやすいページ/印刷/表示設定を補完する。"""
    for attr in (
        "sheet_format",
        "sheet_properties",
        "page_margins",
        "page_setup",
        "print_options",
        "views",
        "row_breaks",
        "col_breaks",
    ):
        _copy_openpyxl_attr(source_ws, target_ws, attr)

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_state = source_ws.sheet_state
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_view.showRowColHeaders = (
        source_ws.sheet_view.showRowColHeaders
    )
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_view.zoomScaleNormal = source_ws.sheet_view.zoomScaleNormal

    target_ws.column_dimensions.clear()
    for key, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key] = copy(dimension)
        target_ws.column_dimensions[key].worksheet = target_ws

    target_ws.row_dimensions.clear()
    for key, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key] = copy(dimension)
        target_ws.row_dimensions[key].worksheet = target_ws

    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.print_title_rows = source_ws.print_title_rows
    target_ws.print_title_cols = source_ws.print_title_cols
    if source_ws.print_area:
        target_ws.print_area = source_ws.print_area


def write_kouten_excel(
    results: list[IntersectionResult],
    template_path: str,
    output_path: str,
    header: KoutenHeaderInfo | None = None,
) -> None:
    """交点計算指示書をExcelに出力する。

    交点をブロック番号ごとにグループ分けし、ブロックごとに
    シートを作成する。各シートは横3ブロック x 縦5段 = 最大15交点。
    15交点を超える場合は同一ブロックで複数シートを作成する。

    Args:
        results: 交点計算結果のリスト
        template_path: テンプレートファイルパス
        output_path: 出力ファイルパス
        header: ヘッダー共通情報（地区名、記入者等）
    """
    if not results:
        raise ValueError("結果データが空です")

    # 出力ディレクトリを作成
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # ブロック番号ごとにグループ分け
    from collections import OrderedDict
    block_groups: OrderedDict[str, list[IntersectionResult]] = OrderedDict()
    for r in results:
        block = _extract_block(r.intersection_stake) or "0"
        if block not in block_groups:
            block_groups[block] = []
        block_groups[block].append(r)

    # テンプレートを開く
    wb = openpyxl.load_workbook(template_path)
    template_ws = wb.active

    for block_num, block_results in block_groups.items():
        # 15交点ごとにページ分割
        pages = (len(block_results) + _KOUTEN_MAX_POINTS - 1) // _KOUTEN_MAX_POINTS

        for page in range(pages):
            page_results = block_results[
                page * _KOUTEN_MAX_POINTS:(page + 1) * _KOUTEN_MAX_POINTS
            ]

            suffix = f"_{page + 1}" if pages > 1 and page > 0 else ""
            sheet_name = f"ブロック{block_num}{suffix}"
            ws = wb.copy_worksheet(template_ws)
            _copy_openpyxl_sheet_layout(template_ws, ws)
            ws.title = sheet_name

            _write_kouten_sheet(ws, page_results, block_number=block_num, header=header)

    wb.remove(template_ws)

    # 保存
    wb.save(output_path)
